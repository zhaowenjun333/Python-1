# !usrbinpython3
# coding=utf-8
import xlwt
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

rb1 = load_workbook('./birds_merchants.xlsx')
rb2 = load_workbook('./merchants.xlsx')

sheet1 = rb1['Sheet1']['A'][1:]
sheet2 = rb2['Sheet1']['A'][1:] + rb2['Sheet1(2)']['A']

cols1 = []
for col1 in sheet1:
    cols1.append(col1.value.split('/')[-1])
print(cols1)

cols2 = []
for col2 in sheet2:
    cols2.append(col2.value)
print(cols2)

book = xlwt.Workbook(encoding='utf-8', style_compression=0)
sheet3 = book.add_sheet('Sheet1', cell_overwrite_ok=True)
sheet3.write(0, 0, 'domain')
filename1 = './new.xlsx'
filename2 = './new.txt'
# cols3 = []
with open(filename2, 'w', encoding='utf-8') as f:
    for i in cols1:
        if i not in cols2:
            f.write(i)
            f.write('\n')
            # cols3.append(i)
            print(i)
        else:
            continue
    f.close()
# for j in range(len(cols3)):
#     sheet3.write(j+1, 0, cols3[j])
#     print(f'{j+1}: {cols3[j]}')
book.save(filename1)

print('save success!')



